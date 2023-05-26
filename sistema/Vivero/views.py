from django.shortcuts import render
from django.urls import reverse, reverse_lazy
from django.http import HttpResponseRedirect, HttpResponse
from django.views.generic import CreateView, TemplateView, ListView, UpdateView, DeleteView
from .models import Vivero
from .forms import DepartamentoMunicipioVivero, FormularioVivero, FormularioMunicipio, FormularioDepartamento
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

#from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
# Create your views here.

#def vivero(request):
#    return render(request, "vivero/viveroFormulario.html")

#@permission_required('is_superuser')
class ViveroCreateView(CreateView):
    model = Vivero
    form_class = FormularioVivero
    template_name = 'vivero/viveroFormulario.html'
    #form_class = FormularioVivero
    #second_form_class = FormularioMunicipio
    #third_form_class = FormularioMunicipio
    #queryset = Vivero.objects.filter()
    success_url = reverse_lazy('vivero:listar_vivero')

    """
    def get_context_data(self, **kwargs):
        context = super(ViveroCreateView, self).get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET)
        return context
    
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST)
        form2 = self.second_form_class(request.POST)

        if form.is_valid() and form2.is_valid():
            vivero = form2.save(commit=False)
    """

    """
    def get_context_data(self, **kwargs):
        context = super(ViveroCreateView, self).get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET)
        if 'form3' not in context:
            context['form3'] = self.third_form_class(self.request.GET)
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST)
        form2 = self.second_form_class(request.POST)
        form3 = self.third_form_class(request.POST)
        if form.is_valid() and form2.is_valid and form3.is_valid:
            departamento = form2.save(commit=False)
            municipio = form3.save(commit=False)
            departamento.municipio = form3.save()
            departamento.vivero = form.save()
            departamento.save()
            return HttpResponseRedirect(self.get_success_url)
        else:
            return self.render_to_response(self.get_context_data(form = form, form2 = form2, form3 = form3))

    """



"""
    def form_valid(self, form):
        departamento = form['departamento'].save()
        municipio = form['municipio'].save(commit = False)
        vivero = form['vivero'].save(commit = False)
        municipio.departamento = departamento
        municipio.save()
        vivero.departamento = departamento
        vivero.save()
        return HttpResponseRedirect(reverse('success'))

class SuccessView(TemplateView):
    template_name = 'vivero/success.html'
"""


class ListarVivero(ListView):
    model = Vivero
    template_name = "vivero/tablaViveros.html"                                                                          

#@permission_required('is_superuser')
class EditarVivero(UpdateView):
    model = Vivero
    template_name = "vivero/viveroFormulario.html"
    form_class = FormularioVivero
    success_url = reverse_lazy('vivero:listar_vivero')

#@permission_required('is_superuser')
class EliminarVivero(DeleteView):
    model = Vivero
    success_url = reverse_lazy('vivero:listar_vivero')

def exportar_viveros_excel(request):
    # Obtener todos los registros de Vivero
    viveros = Vivero.objects.all()

    # Crear un libro de Excel y una hoja de cálculo
    libro_excel = openpyxl.Workbook()
    hoja_excel = libro_excel.active

    # Escribir encabezados de columna en la hoja de cálculo
    hoja_excel['A1'] = 'ID'
    hoja_excel['B1'] = 'Productor'
    hoja_excel['C1'] = 'Departamento'
    hoja_excel['D1'] = 'Municipio'
    hoja_excel['E1'] = 'Código'
    hoja_excel['F1'] = 'Nombre del Vivero'

    # Escribir datos de los viveros en la hoja de cálculo
    for index, vivero in enumerate(viveros, start=2):
        hoja_excel.cell(row=index, column=1).value = vivero.id
        hoja_excel.cell(row=index, column=2).value = vivero.productor.nombre
        hoja_excel.cell(row=index, column=3).value = vivero.departamento
        hoja_excel.cell(row=index, column=4).value = vivero.municipio
        hoja_excel.cell(row=index, column=5).value = vivero.codigo
        hoja_excel.cell(row=index, column=6).value = vivero.nombre_vivero

    # Configurar la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=viveros.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    libro_excel.save(response)

    return response

def exportar_viveros_pdf(request):
    # Obtener todos los registros de Vivero
    viveros = Vivero.objects.all()

    # Crear un archivo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=viveros.pdf'

    # Crear el objeto Canvas del archivo PDF
    pdf = canvas.Canvas(response, pagesize=letter)

    # Definir las posiciones de las columnas
    columnas = [
        (40, 'ID'),
        (100, 'Productor'),
        (200, 'Departamento'),
        (300, 'Municipio'),
        (400, 'Código'),
        (500, 'Nombre del Vivero')
    ]

    # Escribir encabezados de columna en el archivo PDF
    for x, columna in columnas:
        pdf.drawString(x, 750, columna)

    # Escribir datos de los viveros en el archivo PDF
    y = 730  # Posición inicial de las filas
    for vivero in viveros:
        pdf.drawString(40, y, str(vivero.id))
        pdf.drawString(100, y, vivero.productor.nombre)
        pdf.drawString(200, y, vivero.departamento)
        pdf.drawString(300, y, vivero.municipio)
        pdf.drawString(400, y, str(vivero.codigo))
        pdf.drawString(500, y, vivero.nombre_vivero)
        y -= 20  # Espacio entre filas

    # Finalizar el archivo PDF
    pdf.showPage()
    pdf.save()

    return response



