from django.shortcuts import render, redirect
from django.template import loader, Context
from django.http import HttpResponse, HttpResponseRedirect
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
#from django.shortcuts import render_to_response, get_object_or_404
from Productor.forms import FormularioProductor
from Productor.models import Productor
from django.views.generic.edit import CreateView
from django.contrib import messages

from django.contrib.auth.decorators import login_required
# Create your views here.
#@permission_required('is_superuser')
@login_required
def productor(request):
    if request.method == 'POST':
        form = FormularioProductor(request.POST)
        if form.is_valid() and request.user.is_staff:
            #messages.success(request, 'Productor guardado')
            form.save()
        return redirect('Productor')
    else:
        form = FormularioProductor
    return render(request, 'productor/productores.html', {'form': form})


def productor_list(request):
    productor = Productor.objects.all()
    contexto = {'productores': productor}
    return render(request, 'productor/productoresTabla.html', contexto)


def editar_productor(request, id):
    productor = Productor.objects.get(id = id)

    if request.method == 'GET':
        form = FormularioProductor(instance=productor)
        contexto = { 'form': form }
    else:
        form = FormularioProductor(request.POST, instance = productor)
        contexto = { 'form': form }
        if form.is_valid():
            form.save()
            return redirect('listarproductor')
    return render(request, 'productor/productores.html', contexto)


def eliminar_productor(request, id):
    productor = Productor.objects.get(id = id)

    if request.method == 'POST':
        productor.delete()
        return redirect('listarproductor')
    return render(request, 'productor/confirmarEliminar.html', {'productor':productor})
    

def exportar_excel(request):
    # Obtener todos los registros de Productor
    productores = Productor.objects.all()

    # Crear un libro de Excel y una hoja de cálculo
    libro_excel = openpyxl.Workbook()
    hoja_excel = libro_excel.active

    # Escribir encabezados de columna en la hoja de cálculo
    hoja_excel['A1'] = 'Número de Documento'
    hoja_excel['B1'] = 'Nombre'
    hoja_excel['C1'] = 'Apellido'
    hoja_excel['D1'] = 'Dirección'
    hoja_excel['E1'] = 'Email'
    hoja_excel['F1'] = 'Teléfono'

    # Escribir datos de los productores en la hoja de cálculo
    for index, productor in enumerate(productores, start=2):
        hoja_excel.cell(row=index, column=1).value = productor.numero_documento
        hoja_excel.cell(row=index, column=2).value = productor.nombre
        hoja_excel.cell(row=index, column=3).value = productor.apellido
        hoja_excel.cell(row=index, column=4).value = productor.direccion
        hoja_excel.cell(row=index, column=5).value = productor.email
        hoja_excel.cell(row=index, column=6).value = productor.telefono

    # Configurar la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productores.xlsx'

    # Guardar el libro de Excel en la respuesta HTTP
    libro_excel.save(response)

    return response

def exportar_pdf(request):
    # Obtener todos los registros de Productor
    productores = Productor.objects.all()

    # Crear un archivo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=productores.pdf'

    # Crear el objeto Canvas del archivo PDF
    pdf = canvas.Canvas(response, pagesize=letter)

    # Definir las posiciones de las columnas
    columnas = [
        (40, 'Número de Documento'),
        (100, 'Nombre'),
        (200, 'Apellido'),
        (280, 'Dirección'),
        (380, 'Email'),
        (460, 'Teléfono')
    ]

    # Escribir encabezados de columna en el archivo PDF
    for x, columna in columnas:
        pdf.drawString(x, 750, columna)

    # Escribir datos de los productores en el archivo PDF
    y = 730  # Posición inicial de las filas
    for productor in productores:
        pdf.drawString(40, y, str(productor.numero_documento))
        pdf.drawString(100, y, productor.nombre)
        pdf.drawString(200, y, productor.apellido)
        pdf.drawString(280, y, productor.direccion)
        pdf.drawString(380, y, productor.email)
        pdf.drawString(460, y, productor.telefono)
        y -= 20  # Espacio entre filas

    # Finalizar el archivo PDF
    pdf.showPage()
    pdf.save()

    return response